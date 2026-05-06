from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from helpers import (
    BUSINESS_UNITS, GREEN_THRESHOLD, WEEK_HOURS,
    fmt_week, is_holiday_week, max_hours_for_week, month_label,
    traffic_light_status, weeks_between,
)

INK = "1A2332"
PAPER = "F6F2E9"
PAPER_ALT = "EFE9DC"
BORDER = "E3DCCB"
ACCENT = "B8860B"
SAGE = "4A7C59"
OCHRE = "C08F3F"
BRICK = "A0402E"
BRICK_DEEP = "7A2018"
WHITE = "FFFFFF"
SLATE_MUTED = "64748B"

THIN = Side(style="thin", color=BORDER)
HEADER_TOP = Side(style="thin", color=INK)
HEADER_BOTTOM = Side(style="medium", color=ACCENT)


def _generated_at() -> str:
    return datetime.now().strftime("%b %d, %Y · %-I:%M %p")


def _apply_title(ws, title: str, subtitle: str, span: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Calibri", size=18, bold=True, color=INK)
    c.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = Font(name="Calibri", size=10, italic=True, color=SLATE_MUTED)
    c2.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6


def _style_header_row(ws, row_idx: int, count: int) -> None:
    ws.row_dimensions[row_idx].height = 22
    for c in range(1, count + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = Font(name="Calibri", size=11, bold=True, color=PAPER)
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        cell.border = Border(
            top=HEADER_TOP, bottom=HEADER_BOTTOM,
            left=Side(style="thin", color=INK), right=Side(style="thin", color=INK),
        )


def _band_row(ws, row_idx: int, count: int, alt: bool) -> None:
    ws.row_dimensions[row_idx].height = 18
    for c in range(1, count + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = Font(name="Calibri", size=10, color=INK)
        cell.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=PAPER_ALT if alt else WHITE)
        cell.border = Border(bottom=THIN)


def _set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _to_buffer(wb: Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def audit_plan_xlsx(audits, members) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Plan"
    ws.freeze_panes = "A5"

    headers = [
        "Audit", "Phase", "Type", "Business Unit", "Risk", "L", "I",
        "Owner", "Sponsor", "Start", "End", "Budget (h)", "Allocated (h)",
        "% Complete", "Team Size", "Objectives",
    ]
    _set_widths(ws, [44, 14, 16, 18, 12, 4, 4, 18, 18, 10, 10, 12, 12, 12, 11, 60])
    _apply_title(ws, "Audit Plan", f"Generated {_generated_at()} · {len(audits)} audits", len(headers))
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    _style_header_row(ws, 4, len(headers))

    for i, a in enumerate(audits):
        weeks = weeks_between(a.start_week, a.end_week)
        allocated = sum(asgn.hours_per_week * weeks for asgn in a.assignments)
        row = 5 + i
        values = [
            a.name, a.phase, a.type, a.business_unit or "-", a.risk_rating,
            a.likelihood, a.impact, a.owner or "-", a.sponsor or "-",
            fmt_week(a.start_week), fmt_week(a.end_week),
            a.budgeted_hours, allocated, (a.completion_pct or 0) / 100,
            len(a.assignments), a.objectives or "",
        ]
        for c, v in enumerate(values, start=1):
            ws.cell(row=row, column=c, value=v)
        _band_row(ws, row, len(headers), i % 2 == 1)

        risk_cell = ws.cell(row=row, column=5)
        risk_color = (BRICK_DEEP if a.risk_rating == "Critical"
                      else BRICK if a.risk_rating == "High"
                      else OCHRE if a.risk_rating == "Medium"
                      else SAGE)
        risk_cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        risk_cell.fill = PatternFill("solid", fgColor=risk_color)
        risk_cell.alignment = Alignment(vertical="center", horizontal="center")

        if allocated > a.budgeted_hours:
            ws.cell(row=row, column=13).font = Font(name="Calibri", size=10, bold=True, color=BRICK)

        ws.cell(row=row, column=14).number_format = "0%"
        ws.cell(row=row, column=12).number_format = "#,##0"
        ws.cell(row=row, column=13).number_format = "#,##0"
        for c in (6, 7, 14, 15):
            ws.cell(row=row, column=c).alignment = Alignment(vertical="center", horizontal="center")

    # Summary block
    start = ws.max_row + 2
    ws.cell(row=start, column=1, value="Summary").font = Font(name="Calibri", size=12, bold=True, color=INK)
    total_budget = sum(a.budgeted_hours for a in audits)
    total_allocated = sum(
        sum(asgn.hours_per_week * weeks_between(a.start_week, a.end_week) for asgn in a.assignments)
        for a in audits
    )
    summary = [
        ("Total audits", len(audits)),
        ("Active", sum(1 for a in audits if a.phase != "Complete")),
        ("Complete", sum(1 for a in audits if a.phase == "Complete")),
        ("Critical/High risk", sum(1 for a in audits if a.risk_rating in ("Critical", "High"))),
        ("Total budgeted hours", total_budget),
        ("Total allocated hours", total_allocated),
        ("Team size", len(members)),
    ]
    for i, (k, v) in enumerate(summary):
        ws.cell(row=start + 1 + i, column=1, value=k).font = Font(name="Calibri", size=10, color=SLATE_MUTED)
        c = ws.cell(row=start + 1 + i, column=2, value=v)
        c.font = Font(name="Calibri", size=10, bold=True, color=INK)
        c.number_format = "#,##0"

    return _to_buffer(wb)


def team_roster_xlsx(members, audits, member_week_hours, week_keys) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Team Roster"
    ws.freeze_panes = "A5"

    headers = ["Name", "Level", "Email", "This Week", "Annual", "Status", "Assigned Audits"]
    _set_widths(ws, [26, 16, 30, 12, 12, 14, 70])
    _apply_title(ws, "Team Roster", f"Generated {_generated_at()} · {len(members)} members", len(headers))
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    _style_header_row(ws, 4, len(headers))

    current_week = week_keys[0]
    for i, m in enumerate(members):
        row = 5 + i
        this_week = member_week_hours.get(m.id, {}).get(current_week, 0)
        annual = sum(member_week_hours.get(m.id, {}).get(w, 0) for w in week_keys)
        assigned = "; ".join(
            a.name for a in audits if any(asgn.member_id == m.id for asgn in a.assignments)
        )
        status = "Overloaded" if this_week > WEEK_HOURS else "Available" if this_week < GREEN_THRESHOLD else "Utilized"

        values = [m.name, m.level, m.email or "", this_week, annual, status, assigned]
        for c, v in enumerate(values, start=1):
            ws.cell(row=row, column=c, value=v)
        _band_row(ws, row, len(headers), i % 2 == 1)

        sc = ws.cell(row=row, column=6)
        color = BRICK if status == "Overloaded" else SAGE if status == "Available" else OCHRE
        sc.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        sc.fill = PatternFill("solid", fgColor=color)
        sc.alignment = Alignment(vertical="center", horizontal="center")

        ws.cell(row=row, column=4).number_format = "0"
        ws.cell(row=row, column=5).number_format = "#,##0"
        for c in (4, 5):
            ws.cell(row=row, column=c).alignment = Alignment(vertical="center", horizontal="center")

    return _to_buffer(wb)


def utilization_xlsx(members, member_week_hours, week_keys) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "52-Week Utilization"
    ws.freeze_panes = "D6"

    headers = ["Team Member", "Level", "Annual", *[fmt_week(w) for w in week_keys]]
    widths = [26, 16, 12, *[9 for _ in week_keys]]
    _set_widths(ws, widths)
    _apply_title(
        ws, "52-Week Resource Utilization",
        f"Generated {_generated_at()} · {len(members)} members · Starting {fmt_week(week_keys[0])}",
        len(headers),
    )

    # Month band, row 4
    last = ""
    for i, w in enumerate(week_keys):
        ml = month_label(w)
        cell = ws.cell(row=4, column=4 + i, value=ml if ml != last else "")
        cell.fill = PatternFill("solid", fgColor=PAPER_ALT)
        if ml != last:
            cell.font = Font(name="Calibri", size=10, bold=True, color=INK)
            cell.alignment = Alignment(horizontal="left", indent=1)
            last = ml
    ws.row_dimensions[4].height = 18

    # Header row 5
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    _style_header_row(ws, 5, len(headers))
    for i, w in enumerate(week_keys):
        if is_holiday_week(w):
            ws.cell(row=5, column=4 + i, value=f"{fmt_week(w)} ★")
    for i in range(3, len(headers) + 1):
        ws.cell(row=5, column=i).alignment = Alignment(vertical="center", horizontal="center")

    # Member rows
    for mi, m in enumerate(members):
        row = 6 + mi
        annual = sum(member_week_hours.get(m.id, {}).get(w, 0) for w in week_keys)
        ws.cell(row=row, column=1, value=m.name)
        ws.cell(row=row, column=2, value=m.level)
        ws.cell(row=row, column=3, value=annual)
        _band_row(ws, row, len(headers), mi % 2 == 1)
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True, color=INK)
        ws.cell(row=row, column=3).font = Font(name="Calibri", size=10, bold=True, color=INK)
        ws.cell(row=row, column=3).alignment = Alignment(vertical="center", horizontal="center")
        ws.cell(row=row, column=3).number_format = "#,##0"

        for i, w in enumerate(week_keys):
            cell = ws.cell(row=row, column=4 + i)
            h = member_week_hours.get(m.id, {}).get(w, 0)
            mx = max_hours_for_week(w)
            cell.value = h
            cell.number_format = "0;-0;–"
            cell.alignment = Alignment(vertical="center", horizontal="center")

            fg = PAPER_ALT if mi % 2 == 1 else WHITE
            txt = INK
            bold = False
            if h > mx:
                fg, txt, bold = "E6C2BB", BRICK_DEEP, True
            elif h > mx * 0.85:
                fg, txt = "E8D4A8", "8B6423"
            elif h >= GREEN_THRESHOLD:
                fg = "C8D0DD"
            elif h > 0:
                fg = "C8D6CB"
            cell.fill = PatternFill("solid", fgColor=fg)
            cell.font = Font(name="Calibri", size=10, color=txt, bold=bold)

    # Totals row
    totals_row = 6 + len(members)
    week_totals = [
        sum(member_week_hours.get(m.id, {}).get(w, 0) for m in members) for w in week_keys
    ]
    grand_total = sum(week_totals)
    ws.cell(row=totals_row, column=1, value="Team Total")
    ws.cell(row=totals_row, column=2, value="")
    ws.cell(row=totals_row, column=3, value=grand_total)
    for i, t in enumerate(week_totals):
        ws.cell(row=totals_row, column=4 + i, value=t)
    ws.row_dimensions[totals_row].height = 22
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=totals_row, column=c)
        cell.font = Font(name="Calibri", size=11, bold=True, color=PAPER)
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.number_format = "#,##0"
    ws.cell(row=totals_row, column=1).alignment = Alignment(vertical="center", horizontal="left", indent=1)

    # Available row
    avail_row = totals_row + 1
    week_avail = [
        len(members) * max_hours_for_week(w) - week_totals[i] for i, w in enumerate(week_keys)
    ]
    avail_total = sum(max(0, v) for v in week_avail)
    ws.cell(row=avail_row, column=1, value="Available Capacity")
    ws.cell(row=avail_row, column=3, value=avail_total)
    for i, v in enumerate(week_avail):
        ws.cell(row=avail_row, column=4 + i, value=v)
    ws.row_dimensions[avail_row].height = 22
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=avail_row, column=c)
        v = cell.value
        color = BRICK if isinstance(v, (int, float)) and v < 0 else SAGE
        cell.font = Font(name="Calibri", size=10, bold=True, color=color)
        cell.fill = PatternFill("solid", fgColor=PAPER_ALT)
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.number_format = "#,##0"
    ws.cell(row=avail_row, column=1).alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.cell(row=avail_row, column=1).font = Font(name="Calibri", size=11, bold=True, color=INK)

    return _to_buffer(wb)


def executive_xlsx(audits, members) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    ws.freeze_panes = "A5"

    today = datetime.now().date().isoformat()


    headers = ["Status", "Audit", "Business Unit", "Phase", "Risk", "Owner", "% Complete", "End Week"]
    _set_widths(ws, [10, 44, 18, 14, 12, 18, 12, 12])
    _apply_title(ws, "Executive Summary, Audit Committee", f"Generated {_generated_at()} · {len(audits)} engagements · As of {today}", len(headers))
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    _style_header_row(ws, 4, len(headers))

    order = {"Red": 0, "Yellow": 1, "Green": 2}
    ordered = sorted(audits, key=lambda a: (order[traffic_light_status(a)], a.business_unit or ""))

    for i, a in enumerate(ordered):
        row = 5 + i
        status = traffic_light_status(a)
        values = [status, a.name, a.business_unit or "-", a.phase, a.risk_rating, a.owner or "-", (a.completion_pct or 0) / 100, fmt_week(a.end_week)]
        for c, v in enumerate(values, start=1):
            ws.cell(row=row, column=c, value=v)
        _band_row(ws, row, len(headers), i % 2 == 1)
        sc = ws.cell(row=row, column=1)
        color = BRICK if status == "Red" else OCHRE if status == "Yellow" else SAGE
        sc.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        sc.fill = PatternFill("solid", fgColor=color)
        sc.alignment = Alignment(vertical="center", horizontal="center")
        ws.cell(row=row, column=7).number_format = "0%"
        ws.cell(row=row, column=7).alignment = Alignment(vertical="center", horizontal="center")

    # BU breakdown sheet
    ws2 = wb.create_sheet("By Business Unit")
    ws2.freeze_panes = "A5"
    h2 = ["Business Unit", "Total", "Active", "Complete", "Critical/High Risk", "Total Hours"]
    _set_widths(ws2, [22, 10, 10, 12, 18, 14])
    _apply_title(ws2, "Audits by Business Unit", f"Generated {_generated_at()}", len(h2))
    for i, h in enumerate(h2, start=1):
        ws2.cell(row=4, column=i, value=h)
    _style_header_row(ws2, 4, len(h2))

    bu_idx = 0
    for bu in BUSINESS_UNITS:
        bu_audits = [a for a in audits if a.business_unit == bu]
        if not bu_audits:
            continue
        row = 5 + bu_idx
        values = [
            bu, len(bu_audits),
            sum(1 for a in bu_audits if a.phase != "Complete"),
            sum(1 for a in bu_audits if a.phase == "Complete"),
            sum(1 for a in bu_audits if a.risk_rating in ("Critical", "High")),
            sum(a.budgeted_hours for a in bu_audits),
        ]
        for c, v in enumerate(values, start=1):
            ws2.cell(row=row, column=c, value=v)
        _band_row(ws2, row, len(h2), bu_idx % 2 == 1)
        for c in (2, 3, 4, 5, 6):
            ws2.cell(row=row, column=c).alignment = Alignment(vertical="center", horizontal="center")
            ws2.cell(row=row, column=c).number_format = "#,##0"
        bu_idx += 1

    del members
    return _to_buffer(wb)


def activity_log_xlsx(activity) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Log"
    ws.freeze_panes = "A5"

    headers = ["Timestamp", "User", "Action", "Detail"]
    _set_widths(ws, [22, 14, 22, 70])
    _apply_title(ws, "Activity Log", f"Generated {_generated_at()} · {len(activity)} entries", len(headers))
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    _style_header_row(ws, 4, len(headers))

    sorted_log = sorted(activity, key=lambda e: e.timestamp, reverse=True)
    for i, e in enumerate(sorted_log):
        row = 5 + i
        ts = e.timestamp.strftime("%b %d %Y · %-I:%M %p") if hasattr(e.timestamp, "strftime") else str(e.timestamp)
        values = [ts, e.user or "-", e.action, e.detail or ""]
        for c, v in enumerate(values, start=1):
            ws.cell(row=row, column=c, value=v)
        _band_row(ws, row, len(headers), i % 2 == 1)
        ac = ws.cell(row=row, column=3)
        color = BRICK if "Delete" in e.action else SAGE if "Add" in e.action else INK
        ac.font = Font(name="Calibri", size=10, bold=True, color=color)

    return _to_buffer(wb)
